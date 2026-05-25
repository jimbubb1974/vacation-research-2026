import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ─────────────────────────────────────────────────────────────────
HDR_DARK  = "1F3864"
HDR_LIGHT = "D9E1F2"
TOTAL_CLR = "FFE699"
WHITE     = "FFFFFF"
TEXT_LT   = "FFFFFF"

def hdr_font(color=TEXT_LT, bold=True, sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)

def cell_font(bold=False, sz=10):
    return Font(name="Calibri", bold=bold, size=sz)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row, cols, bg=HDR_DARK, fg=TEXT_LT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg)
        cell.font = hdr_font(fg)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin_border()

def write_row(ws, row_num, vals, bg=WHITE, bold=False, num_fmt=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.fill = fill(bg)
        cell.font = cell_font(bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border()
        if num_fmt and c in num_fmt:
            cell.number_format = num_fmt[c]

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# Dollar format
USD = "$#,##0"

# ── Workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# TAB 1 — Summary
# Row map: 2=Greece Opt1, 3=Greece Opt2, 4=Bavaria, 5=StLucia-A, 6=StLucia-B
# Cost/Night formula references C<row> for each destination's ground nights.
# ============================================================
ws = wb.create_sheet("Summary")
ws.row_dimensions[1].height = 36
write_header(ws, 1, [
    "Destination / Option", "Trip Character", "Ground Nights", "Flight Nights",
    "Low Total (USD)", "Mid Total (USD)", "High Total (USD)", "Cost/Night Mid",
    "Cost Confidence", "Logistics Friction", "Weather Risk July",
    "Cultural Depth", "Teen Appeal", "Key Strength", "Key Weakness"
])
summary_rows = [
    # Row 2: Greece Opt 1 — TOTAL row 23 in Greece_Costs
    ["Greece — Opt 1 (Cyclades)", "Ancient city + two-island Cyclades", 9, 1,
     "=Greece_Costs!E23", "=Greece_Costs!F23", "=Greece_Costs!G23", "=Greece_Costs!F23/C2",
     "Medium", "Medium-High", "High (heat + meltemi)", "Highest", "High",
     "Most iconic and memorable trip", "Transfer drag; heat; Santorini crowds"],
    # Row 3: Greece Opt 2 — TOTAL row 21 in Greece2_Costs
    ["Greece — Opt 2 (Crete)", "Crete beaches + history + Athens day", 9, 1,
     "=Greece2_Costs!E21", "=Greece2_Costs!F21", "=Greece2_Costs!G21", "=Greece2_Costs!F21/C3",
     "Medium", "Medium", "Moderate (heat; car on narrow roads)", "Highest", "High",
     "Most historically rich; car freedom; no ferry anxiety",
     "No Santorini caldera view; car adds rental/fuel complexity"],
    # Row 4: Bavaria — TOTAL row 28; Cost/Night uses C4
    ["Bavaria + Salzburg", "Rail-based Alpine and city tour", 9, 1,
     "=Bavaria_Costs!E28", "=Bavaria_Costs!F28", "=Bavaria_Costs!G28", "=Bavaria_Costs!F28/C4",
     "Medium-High", "Medium", "Moderate (mountain weather)", "High", "Very High",
     "Most balanced and stress-light", "Weather-sensitive mountain days; over-design risk"],
    # Row 5: Bavaria Opt 2 — TOTAL row 27 in Bavaria2_Costs
    ["Bavaria — Opt 2 (Cities & Castles)", "Cities and castles (Munich-Berchtesgaden-Salzburg)", 9, 1,
     "=Bavaria2_Costs!E27", "=Bavaria2_Costs!F27", "=Bavaria2_Costs!G27", "=Bavaria2_Costs!F27/C5",
     "Medium-High", "Low-Medium", "Moderate (mountain weather)", "High", "Very High",
     "Neuschwanstein; Hallstatt; castle-focused; less hiking",
     "No alpine village base; Mittenwald removed; similar cost to Opt 1"],
    # Row 6: St. Lucia A — TOTAL row 26; Cost/Night uses C6 (shifted from C5)
    ["St. Lucia — Option A (All-Inclusive)", "Tropical adventure + resort base", 10, 0,
     "=StLucia_Costs!E26", "=StLucia_Costs!F26", "=StLucia_Costs!G26", "=StLucia_Costs!F26/C6",
     "Medium", "Low", "Moderate-High (hurricane season)", "Moderate", "High",
     "Full 10 nights; easy logistics; all-in pricing", "Lower cultural depth; lodging cost dominates"],
    # Row 7: St. Lucia B — TOTAL row 27; Cost/Night uses C7 (shifted from C6)
    ["St. Lucia — Option B (Villa + Bay Gardens)", "Tropical adventure + villa/split", 10, 0,
     "=StLucia_Costs!E27", "=StLucia_Costs!F27", "=StLucia_Costs!G27", "=StLucia_Costs!F27/C7",
     "Medium", "Medium", "Moderate-High (hurricane season)", "Moderate", "High",
     "Best value; most local feel", "One transfer mid-trip; more daily decisions"],
]
d = {5: USD, 6: USD, 7: USD, 8: USD}
for i, row in enumerate(summary_rows):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d)
set_col_widths(ws, [28, 30, 12, 12, 14, 14, 14, 14, 14, 16, 24, 14, 11, 35, 35])
freeze(ws)

# ============================================================
# TAB 2 — Greece Option 1 Itinerary
# ============================================================
ws = wb.create_sheet("Greece_Itinerary")
ws.row_dimensions[1].height = 36
ITIN_COLS = ["Day", "Date", "Location", "Overnight", "Morning", "Afternoon",
             "Evening", "Transport", "Travel Time", "Book Ahead",
             "Weather Sensitive", "Backup Option", "Est. Daily Cost (Family USD)", "Notes"]
write_header(ws, 1, ITIN_COLS)
greece_itin = [
    [1,"Sat Jul 11","DC → Athens","In-flight","Airport departure","Transatlantic flight","In-flight","IAD nonstop or via FRA/IST","Overnight","Y","N","None",0,"Europe clock begins on flight night"],
    [2,"Sun Jul 12","Athens","Athens","Arrive ATH; airport express train; hotel check-in","Plaka walk; Anafiotika lanes; light lunch","Rooftop dinner with Acropolis view","Airport express train","40 min","N","N","Pool/nap day",320,"Keep light; absorb jet lag"],
    [3,"Mon Jul 13","Athens","Athens","Acropolis at 8 am (timed entry required)","Acropolis Museum; shaded lunch; siesta during peak heat","Athens food tour or Monastiraki evening","Local metro/taxi","Local","Y (Acropolis tickets)","Y (heat)","Museum first if very hot",420,"Book Acropolis timed entry well in advance"],
    [4,"Tue Jul 14","Athens → Naxos","Naxos","Blue Star ferry from Piraeus ~07:30","Arrive Naxos ~13:00; check in; Agios Georgios beach","Sunset at the Portara; dinner in Chora","Piraeus ferry","5.5 hrs","Y","Y (wind)","Fly ATH–JNX if seas rough",530,"Blue Star more wind-stable than Seajets; recommended for Piraeus–Naxos"],
    [5,"Wed Jul 15","Naxos","Naxos","Mount Zas hike (start 6:30 am; 3–4 hrs RT; Zeus's Cave en route)","Lunch in Filoti; taxi to Apeiranthos marble village","Halki village; citron distillery tasting; dinner in Chora","Bus or local taxi","Half-day","N","Y (summit heat)","Village loop without the summit hike",320,"Hire taxi for village loop ~€80 half-day"],
    [6,"Thu Jul 16","Naxos","Naxos","Half-day catamaran with snorkeling (departs Chora ~10 am)","Plaka or Agia Anna beach","Waterfront dinner in Naxos Town","Boat; local bus","Full morning","Y","Y (wind cancels boats)","Pure beach day at Plaka if winds strong",520,"Book catamaran 1–2 weeks ahead; best teen day on Naxos"],
    [7,"Fri Jul 17","Naxos → Santorini","Santorini","Late beach swim; lunch in Chora","Seajets fast ferry to Santorini (~1.5 hrs); check in to Imerovigli","Skaros Rock sunset walk; dinner in Imerovigli","Seajets","1.5 hrs","Y","Y (wind cancels Seajets)","Blue Star backup or fly JNX–JTR",840,"Book Seajets early; July Fridays fill fast"],
    [8,"Sat Jul 18","Santorini","Santorini","Fira-to-Oia caldera hike (start 6:30 am; 10 km; ~3.5 hrs)","Oia exploration; lunch; bus/taxi back","Dinner in Imerovigli","Walk + bus/taxi","3.5 hrs hiking","N","N","Sleep in; taxi to Oia instead",580,"Start at dawn; heat arrives fast; single best free activity in Greece"],
    [9,"Sun Jul 19","Santorini","Santorini","Caldera catamaran tour (volcano + hot springs + Red Beach; ~5 hrs)","Akrotiri archaeological site (€15/adult)","Quiet caldera viewpoint dinner","Catamaran + taxi","Full day","Y","Y (boat cancels in wind)","Akrotiri + Red Beach by land if seas rough",860,"Akrotiri requires timed entry; book ahead"],
    [10,"Mon Jul 20","Santorini → Athens","Athens (airport area)","Aegean Airlines morning flight to ATH (~50 min)","Buffer afternoon in Athens; Stavros Niarchos or Syntagma stroll","Early dinner; airport-area hotel (Sofitel Athens Airport)","Domestic flight","50 min + transfer","Y","N","Piraeus fast ferry as backup",490,"Never put international departure at risk of a ferry delay"],
    [11,"Tue Jul 21","Athens → DC","At home","Depart ATH","Transatlantic home","Arrive DC","IAD nonstop or connecting","Long-haul","Y","N","None",0,""],
]
for i, row in enumerate(greece_itin):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt={13: USD})
set_col_widths(ws, [5, 11, 22, 18, 38, 38, 32, 24, 14, 12, 18, 34, 16, 42])
freeze(ws)

# ============================================================
# TAB 3 — Bavaria Itinerary
# ============================================================
ws = wb.create_sheet("Bavaria_Itinerary")
ws.row_dimensions[1].height = 36
write_header(ws, 1, ITIN_COLS)
bavaria_itin = [
    [1,"Sat Jul 11","DC → Munich","In-flight","Airport departure","Overnight transatlantic","In-flight","IAD nonstop (Lufthansa/United)","~8–10 hrs","Y","N","None",0,"Europe clock begins on flight night"],
    [2,"Sun Jul 12","Munich","Munich","Arrive MUC ~9–10 am; S-Bahn to city; hotel check-in","Marienplatz; Viktualienmarkt; Glockenspiel at noon","Englischer Garten + Eisbach surfers; beer garden dinner","S-Bahn S1/S8","40–50 min","N","N","Hotel pool/nap day",300,"Eisbach river surfers free; a genuine teen highlight"],
    [3,"Mon Jul 13","Munich","Munich","Deutsches Museum (4–5 hrs; €15/adult; €4.50/youth)","BMW Welt (free) + BMW Museum (€10/adult)","Hofbräuhaus or neighborhood beer garden","Local U-Bahn","Local","N","N","Extra museum day; Alte Pinakothek alternative",380,"BMW Welt is free and worth 2 hrs even for non-car people"],
    [4,"Tue Jul 14","Munich → Mittenwald","Mittenwald","RB6 train to Mittenwald (1h 50m)","Lüftlmalerei painted houses; Geigenbaumuseum (€5.50)","Karwendelbahn cable car at sunset (optional; free under 15); dinner at Gasthof","Bayern-Ticket + RB6","1h 50m","Y (Bayern-Ticket)","N","Slow afternoon if tired",400,"Bayern-Ticket: 14yo rides free; 17yo = adult; €49/family"],
    [5,"Wed Jul 15","Garmisch day trip","Mittenwald","RB train to Garmisch (~20 min); Partnach Gorge (€7/adult)","Alpspitzbahn panorama (~€30/adult) or summer luge","Return to Mittenwald; dinner","Bayern-Ticket + regional train","20 min each way","Y (Bayern-Ticket)","N","Easy valley walk in Garmisch if crowded",380,"Partnach Gorge is underrated; good in heat"],
    [6,"Thu Jul 16","Zugspitze or Leutaschklamm","Mittenwald","Check zugspitze.de webcam at 7 am. Clear → Zugspitze cogwheel + cable car (~$195 family). Socked in → Leutaschklamm gorge (€5/adult)","Summit/gorge; Eibsee for swim if Zugspitze","Lautersee or Ferchensee lake swim; dinner","Bayern-Ticket + cogwheel train","1.5 hrs to summit from Garmisch","Y (cable car tickets)","Y (Zugspitze critical)","Leutaschklamm + Lautersee lake swim (equally good day)",560,"Never commit to Zugspitze without checking the webcam that morning"],
    [7,"Fri Jul 17","Mittenwald → Berchtesgaden","Berchtesgaden","Mittenwald → Munich (1h 50m) + Munich → Berchtesgaden via Freilassing (2h 30m)","Arrive mid-afternoon; Old Town walk; Royal Palace square","Local dinner; early night","Bayern-Ticket + regional rail","4h 20m total","Y (Bayern-Ticket)","N","Break in Munich midway if needed",350,"Longest luggage day; start ~9 am"],
    [8,"Sat Jul 18","Königssee","Berchtesgaden","Bus 841 to Königssee dock (~10 min); first electric boat to St. Bartholomä — cliff-echo flugelhorn","Hike Salet → Obersee (1.5 km; hidden glacial lake beyond the crowds)","Return by boat; optional salt mine tour (€23/adult)","Bus 841 + electric boat","10 min bus + 35 min each way boat","Y (tickets at dock)","Y (weather-dependent)","Documentation Center Obersalzberg (indoor)",400,"Take first morning boat; crowds build fast after 10 am"],
    [9,"Sun Jul 19","Eagle's Nest + Salzburg","Salzburg","Bus 838; Documentation Center Obersalzberg (€6/adult); Kehlstein bus + tunnel elevator to Eagle's Nest (€36.10/adult; €21/youth; credit card only)","Lunch at Kehlsteinhaus with Watzmann views","Bus 840 Watzmann Express to Salzburg (50 min; ~€11/adult); Old Town evening stroll","Bus 838 + Kehlstein bus + Bus 840","~2 hrs activity + 50 min transfer","Y (Eagle's Nest seasonal)","Y (closes in cloud/wind)","Königssee 2nd day or Ramsau valley hike",510,"Eagle's Nest is historically moving; one of the best teen experiences on the trip"],
    [10,"Mon Jul 20","Salzburg","Salzburg","Hohensalzburg Fortress funicular (€14.50/adult; €8.30/youth)","Mirabell Gardens (free); Mozart's Birthplace (€13.50/adult); Getreidegasse stroll","St. Peter Stiftskeller dinner (Europe's oldest restaurant; book ahead)","Walk + local bus","Local","Y (restaurant)","N","Salzburg Card if doing 3+ paid sites",480,"Compact walkable city; one of the most enjoyable low-effort days on this trip"],
    [11,"Tue Jul 21","Salzburg → Munich → DC","At home","ÖBB Railjet 06:23 or 07:11 Salzburg → Munich (1h 28m; book seat reservation ~$30/pp, 3–4 wks ahead)","S-Bahn S1/S8 to MUC airport (40 min); fly home","Arrive DC","Railjet + S-Bahn","2h 10m total","Y (Railjet reservation)","N","None",170,"Leave Salzburg by 7:15 am for a midday MUC departure"],
]
for i, row in enumerate(bavaria_itin):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt={13: USD})
set_col_widths(ws, [5, 11, 26, 18, 40, 38, 34, 26, 16, 14, 20, 36, 16, 42])
freeze(ws)

# ============================================================
# TAB 4 — St. Lucia Itinerary
# ============================================================
ws = wb.create_sheet("StLucia_Itinerary")
ws.row_dimensions[1].height = 36
write_header(ws, 1, ITIN_COLS)
stlucia_itin = [
    [1,"Sat Jul 11","DC → St. Lucia","Soufrière (B) / Coconut Bay (A)","Drive to PHL; AA nonstop afternoon flight","Arrive UVF ~6–7 pm; transfer to lodging","Welcome dinner; early sleep","PHL→UVF AA nonstop","4h 45m + 50 min transfer","Y","N","Airport-area overnight if very late",320,"Full 10 nights on island; same-day arrival is key advantage over Europe"],
    [2,"Sun Jul 12","Soufrière","Soufrière area","Recovery day: villa pool or resort beach","Light explore of Soufrière town; local lunch","Piton-view dinner at local restaurant","Walk / short taxi","Local","N","Y (modest)","Fully unprogrammed rest day",300,"Protect this day from over-programming; the family needs to settle in"],
    [3,"Mon Jul 13","Soufrière highlights","Soufrière area","Sulphur Springs drive-in volcano (~$10/pp + $5 mud bath)","Toraille Waterfall ($5/pp) + Diamond Botanical Gardens + Mineral Baths ($10/adult)","Local dinner in Soufrière","Half-day driver (~$80 family)","Half-day","Y (driver)","Y (mild)","Swap with Tet Paul Nature Trail if energy low",430,"Hire one driver for all three sites; best all-in-one iconic SL day"],
    [4,"Tue Jul 14","Gros Piton hike","Soufrière area","6:30 am pickup; Gros Piton summit hike (3.5–4 hrs RT; $50/pp trail fee + mandatory guide)","Descend; lunch in Soufrière","Pool recovery; early dinner","Prebooked transfer","45 min to trailhead","Y (guide required)","Y (not in heavy rain)","Tet Paul Nature Trail as lower-intensity alternative",440,"Mandatory guide is a plus — they know the mountain and the flora"],
    [5,"Wed Jul 15","Full-day catamaran","Soufrière area","Sea Spray 'Tout Bagay' or equivalent (~$140/adult; $80/child); departs Rodney Bay or Marigot Bay — Soufrière by sea; Sulphur Springs; Anse Cochon snorkel; open bar","Continue cruise","Return early evening; resort/villa dinner","Prebooked catamaran","Full day","Y (1–2 wks ahead)","Y (seas must be calm)","Land-based Soufrière tour if seas rough",600,"Book early; this is the best teen day on the island"],
    [6,"Thu Jul 16","Rainforest zip-line","Soufrière area","Rainforest Adventures aerial tram + zip-line or Morne Coubaril Estate zip-line with cocoa tour (~$90–95/adult; $75/youth; min age 8; max 290 lbs)","Villa or resort pool; afternoon downtime","Casual dinner; pack for transfer","Prebooked shuttle","~30–45 min","Y","Y (avoid lightning)","Cocoa estate tour only if weather poor",440,"Morne Coubaril combines zip-line with cocoa estate for cultural depth"],
    [7,"Fri Jul 17","Soufrière → Rodney Bay","Rodney Bay area","Late breakfast; optional Tet Paul Nature Trail en route ($10/pp; easy 45-min; best dual-Piton panoramic view)","Transfer north to Bay Gardens Beach Resort (~2 hrs by road or catamaran water taxi)","Anse La Raye Fish Fry (local village; grilled lobster; breadfruit; rum punch; best food experience on island)","Private taxi or catamaran water taxi","1.5–2 hrs","Y (resort; Fish Fry)","Y (mild)","Skip Tet Paul; go straight north",400,"Anse La Raye Fish Fry runs Friday and Saturday; arrange return transport ($30–40)"],
    [8,"Sat Jul 18","Rodney Bay / Gros Islet","Rodney Bay area","Reduit Beach morning","Pigeon Island National Park ($11/adult; Fort Rodney ruins; 18th-century cannons)","Jambe de Bois café at Pigeon Island; marina sundowners","Short taxi","15 min each way","N","Y (mild)","Extra beach day at Reduit",300,"Pigeon Island is underrated; teens like the fort ruins"],
    [9,"Sun Jul 19","Snorkel + downtime","Rodney Bay area","Anse Chastanet snorkel (boat transfer ~$40/pp; turtles; parrotfish; reef drops off beach)","Afternoon beach club; sunbeds; beach bar","Sunset dinner at marina restaurant","Boat or taxi","30–45 min to Anse Chastanet","Optional","Y (visibility-dependent)","Pure resort beach day",380,"Anse Chastanet reef is worth the transfer"],
    [10,"Mon Jul 20","Reserve / flexible","Rodney Bay area","Weather makeup day or pure downtime — pre-book nothing","Shopping; final swim; Castries market visit","Farewell dinner; pack","Local","Local","N","Y","Leave unplanned by design",270,"Most important day to protect from pre-programming in July hurricane season"],
    [11,"Tue Jul 21","St. Lucia → DC","At home","Resort shuttle to UVF (~1.5 hrs from Rodney Bay)","AA nonstop UVF → PHL (~4h 45m); arrive ~7 pm; drive home","At home","Taxi + AA nonstop","1.5 hrs + 4h 45m","Y","N","None",0,""],
]
for i, row in enumerate(stlucia_itin):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt={13: USD})
set_col_widths(ws, [5, 11, 24, 22, 42, 40, 36, 26, 16, 12, 18, 36, 16, 44])
freeze(ws)

# ============================================================
# TAB 5 — Bavaria Option 2 Itinerary (Cities & Castles)
# ============================================================
ws = wb.create_sheet("Bavaria2_Itinerary")
ws.row_dimensions[1].height = 36
write_header(ws, 1, ITIN_COLS)
bavaria2_itin = [
    [1,"Sat Jul 11","DC → Munich","In-flight","Airport departure","Overnight transatlantic","In-flight","IAD nonstop (Lufthansa/United)","~8–10 hrs","Y","N","None",0,"Europe clock begins on flight night"],
    [2,"Sun Jul 12","Munich","Munich","Arrive MUC ~9–10 am; S-Bahn to city; hotel check-in","Marienplatz; Viktualienmarkt; Glockenspiel at noon","Englischer Garten + Eisbach surfers; beer garden dinner","S-Bahn S1/S8","40–50 min","N","N","Hotel pool/nap day",300,"Eisbach river surfers free; a genuine teen highlight"],
    [3,"Mon Jul 13","Munich","Munich","Deutsches Museum (4–5 hrs; €15/adult; €4.50/youth)","BMW Welt (free) + BMW Museum (€10/adult)","Hofbräuhaus or neighborhood beer garden","Local U-Bahn","Local","N","N","Extra museum day; Alte Pinakothek alternative",380,"BMW Welt is free and worth 2 hrs even for non-car people"],
    [4,"Tue Jul 14","Neuschwanstein day trip","Munich","Regional train to Füssen (~2 hrs); bus to castle; timed interior tour","Marienbrücke bridge view over gorge; optional Hohenschwangau Castle","Return to Munich; dinner","Bayern-Ticket + regional train","2 hrs each way","Y (Neuschwanstein timed entry — book immediately)","N","Exterior + Marienbrücke if interior sells out",430,"Book timed entry at tickets.hohenschwangau.de immediately; July sells out weeks ahead; under 18 free"],
    [5,"Wed Jul 15","Munich","Munich","Nymphenburg Palace + formal gardens (~2 hrs; €15/adult)","Olympic Park grounds (free) + Maximilianstrasse or Schwabing district","Dinner in Munich","U-Bahn + local","Local","N","N","Dachau Memorial if family interested in WWII history",350,"Nymphenburg is a genuine royal palace; canals and formal gardens free even without entry"],
    [6,"Thu Jul 16","Munich → Berchtesgaden","Berchtesgaden","Check out; regional train to Berchtesgaden (~2.5 hrs via Freilassing)","Arrive mid-afternoon; Berchtesgaden Old Town walk; Royal Palace square","Dinner at local Gasthof; early night","Bayern-Ticket + regional rail","~2h 30m","Y (Bayern-Ticket)","N","Break in Munich midway if needed",350,"Bayern-Ticket covers the full day; straightforward connection"],
    [7,"Fri Jul 17","Königssee","Berchtesgaden","Bus 841 to Königssee dock (~10 min); first electric boat to St. Bartholomä — cliff-echo flugelhorn","Return by boat; afternoon rest or Old Town walk","Dinner in Berchtesgaden","Bus 841 + electric boat","10 min bus + 35 min each way boat","Y (boat dock)","Y (weather-dependent)","Documentation Center Obersalzberg (indoor)",400,"Take first morning boat; crowds build fast after 10 am"],
    [8,"Sat Jul 18","Berchtesgaden → Salzburg","Salzburg","Bus 838; Documentation Center Obersalzberg (€6/adult); Kehlstein bus + tunnel elevator to Eagle's Nest (€36.10/adult; €21/youth)","Berchtesgaden salt mine (~1.5 hrs; €23/adult) — core plan, not optional","Bus 840 to Salzburg (~50 min); check in; light dinner","Bus 838 + Bus 840","~2 hrs activity + 50 min transfer","Y (Eagle's Nest seasonal)","Y (Eagle's Nest closes in cloud/wind)","Salt mine only + travel to Salzburg if Eagle's Nest closed",520,"Eagle's Nest historically moving; salt mine adds variety; Day 8 is full but manageable"],
    [9,"Sun Jul 19","Salzburg","Salzburg","Hohensalzburg Fortress funicular (€14.50/adult; €8.30/youth)","Mirabell Gardens (free); Mozart's Birthplace (€13.50/adult); Getreidegasse stroll","St. Peter Stiftskeller dinner (Europe's oldest restaurant; book ahead)","Walk + local bus","Local","Y (restaurant reservation)","N","Salzburg Card if doing 3+ paid sites",480,"Compact walkable city; one of the most enjoyable low-effort days on this trip"],
    [10,"Mon Jul 20","Hallstatt day trip","Salzburg","Regional train + bus to Hallstatt (~1.5 hrs each way); lake village walk; boat across the Hallstätter See","Hallstatt salt mine museum (~1 hr); Beinhaus charnel house (unusual and memorable); lakeside lunch","Return to Salzburg; evening Salzach riverside walk","Regional train + bus","1.5 hrs each way","N","N","Salzburg Museum der Moderne or Sound of Music tour instead",400,"Arrive early to beat tour buses; Hallstatt is UNESCO-listed and unmistakably beautiful"],
    [11,"Tue Jul 21","Salzburg → Munich → DC","At home","ÖBB Railjet 06:23 or 07:11 Salzburg → Munich (1h 28m; book seat reservation ~$30/pp, 3–4 wks ahead)","S-Bahn S1/S8 to MUC airport (40 min); fly home","Arrive DC","Railjet + S-Bahn","2h 10m total","Y (Railjet reservation)","N","None",170,"Leave Salzburg by 7:15 am for a midday MUC departure"],
]
for i, row in enumerate(bavaria2_itin):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt={13: USD})
set_col_widths(ws, [5, 11, 26, 18, 40, 40, 34, 26, 16, 22, 20, 36, 16, 44])
freeze(ws)

# ============================================================
# TAB 6 — Greece Option 2 Itinerary (Crete + Athens)
# ============================================================
ws = wb.create_sheet("Greece2_Itinerary")
ws.row_dimensions[1].height = 36
write_header(ws, 1, ITIN_COLS)
greece2_itin = [
    [1,"Sat Jul 11","DC → Athens","In-flight","Airport departure","Transatlantic flight","In-flight","IAD nonstop or via FRA/IST","Overnight","Y","N","None",0,"Europe clock begins on flight night"],
    [2,"Sun Jul 12","Athens → Chania","Chania area","Arrive ATH; connect to CHQ (Aegean ~50 min)","Pick up rental car (CHQ airport); check in to Chania beach hotel","Easy beach swim; dinner near hotel","ATH→CHQ domestic + rental car pickup","~50 min flight + 20 min to hotel","Y (Aegean domestic)","N","Relax at hotel on delayed arrival",360,"Agia Marina / Stalos area; beach within walking distance of hotel"],
    [3,"Mon Jul 13","Chania area","Chania area","Recovery beach morning; pool time","Pool or beach; light lunch near hotel","Chania Old Town + Venetian Harbor evening walk","Short drive (~15 min)","Local","N","N","Full beach day",300,"Chania Old Town is walkable; Venetian lighthouse is free"],
    [4,"Tue Jul 14","West Crete","Chania area","Balos Lagoon (early start) — Kissamos ferry (~1.5 hrs) or drive to overlook","Beach / swim at Balos or Falassarna (nearby alternative)","Casual seafood dinner; return to hotel","Rental car + Kissamos ferry","~1.5 hrs each way ferry","Y (Kissamos ferry)","Y (ferry cancels in wind)","Falassarna beach by car — equally scenic; no ferry",480,"Balos ferry fills in July — pre-book Kissamos ferry tickets"],
    [5,"Wed Jul 15","Chania / villages","Chania area","Chania Old Town covered market + municipal market; morning shop","Inland village drive (Vamos or Gavalochori); olive oil tasting","Dinner in Chania","Rental car","~30 min to villages","N","N","Extra beach morning instead",280,"Cretan olive oil tasting ~€20/pp; village kafeneion lunch is good value"],
    [6,"Thu Jul 16","Imbros Gorge","Chania area","Imbros Gorge hike early start (~7 am; 8 km point-to-point; ~3.5 hrs)","South coast lunch in Sfakia; swim at Sweetwater Beach (short boat from Sfakia)","Low-key evening; early dinner; pack for drive east tomorrow","Rental car (~45 min to trailhead)","45 min drive; 3.5 hr hike","N","Y (heat — start early)","Beach day at Frangokastello instead",380,"Gorge entry ~€5/pp; arrange taxi back from Sfakia (~€15 total)"],
    [7,"Fri Jul 17","Chania → East Crete","East Crete","Check out; leisurely drive east on E75 (2.5–3 hrs total)","Stop Rethymno Old Town — Venetian fortress (Fortezza), harbor lunch (~1 hr)","Check in Agios Nikolaos; seaside harbor dinner","Rental car (~2.5–3 hrs with stop)","2.5–3 hrs driving","N","N","Skip Rethymno if tired; drive straight",400,"Rethymno Fortezza ~€5 entry; worth 30–45 min; harbor seafood lunch"],
    [8,"Sat Jul 18","East Crete — mythology day","East Crete","Knossos Palace early (opens 8 am; 1.5–2 hrs; book timed entry)","Heraklion Archaeological Museum (2–3 hrs; excellent A/C midday)","Heraklion harbor dinner; return to Agios Nikolaos (~1 hr)","Rental car (~1 hr to Heraklion)","~1 hr each way","Y (Knossos timed entry)","Y (heat)","Museum only if exhausted",490,"Knossos + museum make the strongest possible Minoan double; don't split"],
    [9,"Sun Jul 19","East Crete — island day","East Crete","Spinalonga boat trip from Elounda (~20 min crossing; 1.5 hrs on island)","Beach or pool afternoon in Elounda","Elounda or Agios Nikolaos harbor dinner","Rental car + local boat","20 min boat each way","Y (morning boat)","Y (boat cancels in wind)","Agios Nikolaos Archaeological Museum if seas rough",420,"Spinalonga leper colony island is historically striking; book early boat to avoid crowds"],
    [10,"Mon Jul 20","Crete → Athens","Athens","Return rental car at HER airport; fly HER→ATH (Aegean ~45 min)","Check in Athens (Plaka or Koukaki area); afternoon rest","Plaka walk + rooftop dinner with Acropolis view","Rental car → HER + domestic flight","~45 min flight","Y (Aegean; rental return timing)","N","Airport hotel if very late arrival",480,"Return car before check-in to avoid Athens city parking fees"],
    [11,"Tue Jul 21","Athens → DC","At home","Acropolis timed entry 7:30–9 am (FIRM; pre-book) OR relaxed hotel breakfast if flight too early","Fly ATH→IAD (depart midday or afternoon)","Arrive DC","Taxi to ATH airport + transatlantic","Long-haul","Y (Acropolis — must pre-book)","N","Skip Acropolis if flight is early morning",170,"Book Acropolis timed entry immediately; July 7:30 am slots sell out weeks ahead"],
]
for i, row in enumerate(greece2_itin):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt={13: USD})
set_col_widths(ws, [5, 11, 26, 18, 40, 38, 34, 26, 16, 14, 20, 36, 16, 44])
freeze(ws)

# ============================================================
# TAB 6 — Greece Option 1 Costs
# Row map (data rows 2-22, TOTAL row 23):
#   2-3: Flights, Bags
#   4-7: Lodging (Athens-2n, Naxos-3n, Santorini-3n, Athens-buf)
#   8-12: Ferries + Internal flight + Transfers
#   13-20: Activities (8 rows)
#   21: Food
#   22: Insurance
# ============================================================
ws = wb.create_sheet("Greece_Costs")
ws.row_dimensions[1].height = 36
COST_COLS = ["Category", "Item", "Unit Cost", "Qty",
             "Low (USD)", "Mid (USD)", "High (USD)",
             "Confidence", "Basis", "Source IDs"]
write_header(ws, 1, COST_COLS)

greece_costs = [
    # row 2
    ["Flights","IAD→ATH round-trip economy × 4","$1,100–1,600/pp","4",4000,5200,6400,"Medium","Live IAD–ATH searches May 2026; United nonstop ~$1,350–1,467 observed","GR-FL"],
    # row 3
    ["Bags & Seats","Economy bags + seat selection","$90/pp RT","4",200,360,560,"Low","Carrier-dependent; assume 1 checked bag/pp","GR-FL"],
    # row 4
    ["Lodging","Athens — 2 nights (arrival)","$200–350/night","2",400,480,700,"Medium","Plaka-area hotels (Athens Gate, Electra Metropolis range)","GR-LO"],
    # row 5
    ["Lodging","Naxos — 3 nights","$150–375/night","3",450,750,1125,"Medium","Iria Beach Art Hotel or similar mid-range","GR-LO"],
    # row 6
    ["Lodging","Santorini — 3 nights","$300–800/night","3",900,1500,2400,"Medium","Imerovigli/Firostefani caldera-adjacent; Oia center commands premium","GR-LO"],
    # row 7
    ["Lodging","Athens buffer — 1 night (airport area)","$180–280/night","1",180,200,280,"Medium","Sofitel Athens Airport or similar","GR-LO"],
    # row 8
    ["Ferries","Piraeus → Naxos (Blue Star mid or Seajets)","$55–100/pp","4",220,300,400,"Medium-High","Blue Star Ferries and Seajets direct schedules","GR-FE"],
    # row 9
    ["Ferries","Naxos → Santorini (Seajets fast ferry)","$65–90/pp","4",180,260,360,"Medium-High","Seajets Naxos–Santorini May 2026 fares","GR-FE"],
    # row 10
    ["Internal Flight","Santorini → Athens (Aegean Airlines domestic)","$70–130/pp","4",0,480,560,"Medium","Aegean Airlines; optional; ferry is the budget path back","GR-DF"],
    # row 11
    ["Transfers","Athens airport express train (round-trip)","$18/pp RT","4",40,72,72,"High","Proastiakos line; under-18 half price","GR-TR"],
    # row 12
    ["Transfers","Island taxis, buses, water taxis","Estimate","—",200,350,500,"Low","Naxos buses cheap; Santorini taxis expensive in July","GR-TR"],
    # row 13
    ["Activities","Acropolis timed entry (advance)","€30/adult; €15/youth","4",120,170,170,"High","Greek Ministry of Culture / hhticket.gr; €30 adult flat rate 2026","GR-AC"],
    # row 14
    ["Activities","Acropolis Museum","€15/adult","4",60,71,71,"High","Official Acropolis Museum site","GR-AC"],
    # row 15
    ["Activities","Ancient Agora","€10/adult","4",0,47,47,"High","Greek Ministry of Culture","GR-AC"],
    # row 16
    ["Activities","Naxos catamaran half-day","€65–90/pp","4",200,280,400,"Medium","Operator pricing estimated; varies by provider","GR-AC"],
    # row 17
    ["Activities","Santorini caldera catamaran (volcano + hot springs + Red Beach)","€100–130/pp","4",320,440,590,"Medium","Operator estimates; book through Viator or direct","GR-AC"],
    # row 18
    ["Activities","Akrotiri archaeological site","€15/adult","4",30,71,71,"High","Greek Ministry of Culture","GR-AC"],
    # row 19
    ["Activities","Athens food/walking tour (optional)","$55–70/pp","4",0,220,280,"Low","Tour operator estimates","GR-AC"],
    # row 20
    ["Activities","Misc (Portara free; Naxos villages; incidentals)","Estimate","—",0,60,120,"Low","Estimate","—"],
    # row 21
    ["Food & Dining","9 ground days × 4 people","$50/$80/$120 pp/day","36 pp-days",1800,2880,4320,"Medium","Mix of taverna lunches ($15–25/pp) + dinners ($30–60/pp); islands more expensive than Athens","GR-FD"],
    # row 22
    ["Travel Insurance","Trip protection (ferries + weather)","$80/pp","4",200,320,500,"Medium","Recommended; covers ferry delays","—"],
]

d3 = {5: USD, 6: USD, 7: USD}
for i, row in enumerate(greece_costs):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d3)

# TOTAL row 23 — formulas sum rows 2:22
write_row(ws, 23,
    ["TOTAL", "", "", "", "=SUM(E2:E22)", "=SUM(F2:F22)", "=SUM(G2:G22)", "", "", ""],
    bg=TOTAL_CLR, bold=True, num_fmt=d3)

set_col_widths(ws, [18, 44, 22, 10, 12, 12, 12, 14, 50, 12])
freeze(ws)

# ============================================================
# TAB 7 — Bavaria Costs
# Row map (data rows 2-27, TOTAL row 28):
#   2-3: Flights, Bags
#   4-7: Lodging (Munich-2n, Mittenwald-3n, Berchtesgaden-2n, Salzburg-2n)
#   8-12: Rail + Transit
#   13-25: Activities (13 rows)
#   26: Food
#   27: Insurance
# ============================================================
ws = wb.create_sheet("Bavaria_Costs")
ws.row_dimensions[1].height = 36
write_header(ws, 1, COST_COLS)

bavaria_costs = [
    # row 2
    ["Flights","IAD→MUC round-trip economy × 4","$950–1,600/pp","4",3800,4800,6400,"Medium-High","Lufthansa and United IAD–MUC nonstop ~8h 10m; Icelandair via KEF lower","BA-FL"],
    # row 3
    ["Bags & Seats","Economy bags + seat selection","$90/pp RT","4",200,300,450,"Low","Carrier-dependent","BA-FL"],
    # row 4
    ["Lodging","Munich — 2 nights","$200–450/night","2",400,600,900,"Medium","Torbräu / NH Collection / Eurostars Grand Central range","BA-LO"],
    # row 5
    ["Lodging","Mittenwald — 3 nights","$180–400/night","3",540,810,1200,"Medium","Alpenhotel Rieger / Post Hotel range","BA-LO"],
    # row 6
    ["Lodging","Berchtesgaden — 2 nights","$180–450/night","2",360,560,900,"Medium","Hotel Edelweiss / Vier Jahreszeiten range","BA-LO"],
    # row 7
    ["Lodging","Salzburg — 2 nights (Festspiele premium applies)","$250–600/night","2",500,750,1200,"Medium","Hotel Wolf-Dietrich; book 5–6 months ahead for Festival dates","BA-LO"],
    # row 8
    ["Rail","Bayern-Ticket family × 4 days (14yo free; 17yo = adult; €49/day)","€49/day","4 days",220,231,231,"High","DB official Bayern-Ticket; children 6–14 ride free with adult","BA-RT"],
    # row 9
    ["Rail","Salzburg → Munich Railjet + seat reservation","~$35/pp","4",100,140,200,"High","ÖBB Railjet; book 3–4 weeks ahead with seat reservation","BA-RT"],
    # row 10
    ["Rail","MUC airport S-Bahn round-trip (2 × family day ticket)","$70/family × 2","2 trips",80,140,140,"High","MVV airport day ticket","BA-RT"],
    # row 11
    ["Transit","Bus 840 Berchtesgaden → Salzburg (family)","~€11/pp","4",35,52,52,"High","Regional bus; not covered by Bayern-Ticket","BA-RT"],
    # row 12
    ["Transit","Local Berchtesgaden buses (Eagle's Nest + Königssee; 2 days)","~€8/pp/day","4 pp × 2 days",30,38,38,"High","Berchtesgaden local transit","BA-RT"],
    # row 13
    ["Activities","Deutsches Museum","€15/adult; €4.50/youth","2+2",30,45,45,"High","Deutsches Museum official site","BA-AC"],
    # row 14
    ["Activities","BMW Welt (free) + BMW Museum","€10/adult","2",0,24,24,"High","BMW Museum official site","BA-AC"],
    # row 15
    ["Activities","Karwendelbahn cable car (3 adults incl. 17yo; 14yo free)","€33/adult","3",0,117,117,"High","Karwendelbahn official; free under 15","BA-AC"],
    # row 16
    ["Activities","Zugspitze family combo (cogwheel + cable car)","~$195 family","1",180,195,230,"Medium","Zugspitze.de summer 2026; family/youth pricing applies","BA-AC"],
    # row 17
    ["Activities","Partnach Gorge","€7/adult","4",30,33,33,"High","Official Partnach Gorge entry","BA-AC"],
    # row 18
    ["Activities","Königssee electric boat round-trip to Salet","€22/adult; €11/youth (14yo)","4",80,92,92,"High","Bayerische Seenschifffahrt official pricing","BA-AC"],
    # row 19
    ["Activities","Eagle's Nest (Kehlstein bus + tunnel elevator)","€36.10/adult; €21/youth","4",110,143,143,"High","Official Kehlstein; credit card only at top","BA-AC"],
    # row 20
    ["Activities","Documentation Center Obersalzberg","€6/adult","4",28,28,28,"High","Official Berchtesgaden tourism","BA-AC"],
    # row 21
    ["Activities","Hohensalzburg Fortress funicular","€14.50/adult; €8.30/youth","4",45,63,63,"High","Festung Salzburg official","BA-AC"],
    # row 22
    ["Activities","Mozart's Birthplace","€13.50/adult","4",0,64,64,"High","Mozarteum Foundation","BA-AC"],
    # row 23
    ["Activities","Optional: Neuschwanstein guided day tour","~$75/pp","4",0,0,300,"Low","Guided tour from Munich; NOT in core plan","BA-AC"],
    # row 24
    ["Activities","Optional: Berchtesgaden salt mine","€23/adult","4",0,0,109,"Medium","Salzbergwerk official site","BA-AC"],
    # row 25
    ["Activities","Misc (village fees; incidentals)","Estimate","—",50,80,120,"Low","Estimate","—"],
    # row 26
    ["Food & Dining","9 ground days × 4 people","$40/$65/$100 pp/day","36 pp-days",1440,2340,3600,"Medium","Bakery breakfasts; beer-garden lunches; one sit-down dinner/day","BA-FD"],
    # row 27
    ["Travel Insurance","Trip protection","$75/pp","4",200,300,450,"Medium","Standard comprehensive policy","—"],
]

for i, row in enumerate(bavaria_costs):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d3)

# TOTAL row 28 — formulas sum rows 2:27
write_row(ws, 28,
    ["TOTAL", "", "", "", "=SUM(E2:E27)", "=SUM(F2:F27)", "=SUM(G2:G27)", "", "", ""],
    bg=TOTAL_CLR, bold=True, num_fmt=d3)

set_col_widths(ws, [18, 46, 24, 14, 12, 12, 12, 14, 52, 12])
freeze(ws)

# ============================================================
# TAB 8 — Bavaria Option 2 Costs
# Row map (data rows 2-26, TOTAL row 27):
#   2-3:   Flights, Bags
#   4-6:   Lodging (Munich 4n, Berchtesgaden 2n, Salzburg 3n)
#   7-11:  Rail + Transit
#   12-24: Activities (13 rows)
#   25:    Food
#   26:    Insurance
# Verified totals: Low $8,644 / Mid $12,309 / High $17,334
# ============================================================
ws = wb.create_sheet("Bavaria2_Costs")
ws.row_dimensions[1].height = 36
write_header(ws, 1, COST_COLS)

bavaria2_costs = [
    # row 2
    ["Flights","IAD→MUC round-trip economy × 4","$950–1,600/pp","4",3800,4800,6400,"Medium-High","Same IAD–MUC routing as Option 1; Lufthansa/United nonstop","BA-FL"],
    # row 3
    ["Bags & Seats","Economy bags + seat selection","$90/pp RT","4",200,300,450,"Low","Carrier-dependent","BA-FL"],
    # row 4
    ["Lodging","Munich — 4 nights","$200–450/night","4",800,1200,1800,"Medium","Same hotel range as Option 1; 2 additional nights","BA-LO"],
    # row 5
    ["Lodging","Berchtesgaden — 2 nights","$180–450/night","2",360,560,900,"Medium","Hotel Edelweiss / Vier Jahreszeiten range","BA-LO"],
    # row 6
    ["Lodging","Salzburg — 3 nights (Festspiele premium applies)","$250–600/night","3",750,1125,1800,"Medium","1 additional night vs. Option 1; book 5–6 months ahead","BA-LO"],
    # row 7
    ["Rail","Bayern-Ticket × 2 days (Neuschwanstein + Munich→Berchtesgaden)","€49/day","2 days",110,116,116,"High","DB Bayern-Ticket; 14yo free; 17yo adult rate","BA-RT"],
    # row 8
    ["Transit","Munich local transit — non–Bayern-Ticket days (MVV day tickets)","~$25/family/day","4 days",60,100,100,"High","MVV day tickets; S-Bahn and U-Bahn","BA-RT"],
    # row 9
    ["Rail","Salzburg → Munich Railjet + seat reservation","~$35/pp","4",100,140,200,"High","ÖBB Railjet; book 3–4 weeks ahead","BA-RT"],
    # row 10
    ["Rail","MUC airport S-Bahn round-trip (2 × family day ticket)","$70/family × 2","2 trips",80,140,140,"High","MVV airport day ticket","BA-RT"],
    # row 11
    ["Transit","Bus 840 Berchtesgaden → Salzburg (family)","~€11/pp","4",35,52,52,"High","Regional bus; not covered by Bayern-Ticket","BA-RT"],
    # row 12
    ["Transit","Local Berchtesgaden buses (Eagle's Nest + Königssee; 2 days)","~€8/pp/day","4 pp × 2 days",30,38,38,"High","Berchtesgaden local transit","BA-RT"],
    # row 13
    ["Activities","Deutsches Museum","€15/adult; €4.50/youth","2+2",30,45,45,"High","Deutsches Museum official site","BA-AC"],
    # row 14
    ["Activities","BMW Welt (free) + BMW Museum","€10/adult","2",0,24,24,"High","BMW Museum official site","BA-AC"],
    # row 15
    ["Activities","Neuschwanstein Castle (timed interior tour + Füssen bus)","€15/adult; under 18 free","4",80,200,280,"Medium","tickets.hohenschwangau.de; under 18 free with guardian; July sells out weeks ahead","BA2-AC"],
    # row 16
    ["Activities","Nymphenburg Palace","€15/adult","4",36,50,50,"High","Schloss Nymphenburg official; grounds free","BA2-AC"],
    # row 17
    ["Activities","Königssee electric boat (round-trip to Salet)","€22/adult; €11/youth (14yo)","4",80,92,92,"High","Bayerische Seenschifffahrt official pricing","BA-AC"],
    # row 18
    ["Activities","Eagle's Nest (Kehlstein bus + tunnel elevator)","€36.10/adult; €21/youth","4",110,143,143,"High","Official Kehlstein; credit card only at top; seasonal","BA-AC"],
    # row 19
    ["Activities","Documentation Center Obersalzberg","€6/adult","4",28,28,28,"High","Official Berchtesgaden tourism","BA-AC"],
    # row 20
    ["Activities","Berchtesgaden salt mine (core plan)","€23/adult + youth est.","4",90,109,109,"Medium","Salzbergwerk official site; now in core plan not optional","BA-AC"],
    # row 21
    ["Activities","Hohensalzburg Fortress funicular","€14.50/adult; €8.30/youth","4",45,63,63,"High","Festung Salzburg official","BA-AC"],
    # row 22
    ["Activities","Mozart's Birthplace","€13.50/adult","4",0,64,64,"High","Mozarteum Foundation","BA-AC"],
    # row 23
    ["Activities","Hallstatt day trip (regional train + bus + salt mine museum entry)","Estimate","4",120,180,240,"Medium","Regional rail Salzburg–Hallstatt; Hallstatt Salzwelten ~€15/adult","BA2-AC"],
    # row 24
    ["Activities","Misc (Olympic Park; village fees; incidentals)","Estimate","—",60,100,150,"Low","Estimate","—"],
    # row 25
    ["Food & Dining","9 ground days × 4 people","$40/$65/$100 pp/day","36 pp-days",1440,2340,3600,"Medium","Same food model as Option 1; bakeries + beer gardens","BA-FD"],
    # row 26
    ["Travel Insurance","Trip protection","$75/pp","4",200,300,450,"Medium","Standard comprehensive policy","—"],
]

for i, row in enumerate(bavaria2_costs):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d3)

# TOTAL row 27 — formulas sum rows 2:26
# Low $8,644 / Mid $12,309 / High $17,334
write_row(ws, 27,
    ["TOTAL", "", "", "", "=SUM(E2:E26)", "=SUM(F2:F26)", "=SUM(G2:G26)", "", "", ""],
    bg=TOTAL_CLR, bold=True, num_fmt=d3)

set_col_widths(ws, [18, 46, 24, 14, 12, 12, 12, 14, 52, 12])
freeze(ws)

# ============================================================
# TAB 10 — St. Lucia Costs (Options A & B combined)
# Row map (data rows 2-25, TOTAL A row 26, TOTAL B row 27):
#   2-3:   Flights, Bags  (Both)
#   4:     Lodging_A Coconut Bay  (Option A only)
#   5-6:   Lodging_B Soufrière villa, Bay Gardens  (Option B only)
#   7-8:   Transfer_A  (Option A only)
#   9-12:  Transfer_B  (Option B only)
#   13-21: Activities — all marked Both
#   22:    Food_A  (Option A only)
#   23:    Food_B  (Option B only)
#   24-25: Tips & Misc, Insurance  (Both)
# TOTAL rows use SUMIF on col K ("Applies To")
# ============================================================
ws = wb.create_sheet("StLucia_Costs")
ws.row_dimensions[1].height = 36
SL_COLS = COST_COLS + ["Applies To"]
write_header(ws, 1, SL_COLS)

stlucia_costs = [
    # row 2
    ["Flights","PHL→UVF round-trip economy × 4 (AA nonstop)","$550–1,000/pp","4",2200,3000,4000,"Medium","AA PHL–UVF weekly nonstop; DC airports add 1 stop + ~$100–150/pp","SL-FL","Both"],
    # row 3
    ["Bags & Seats","Typical fees","$60/pp RT","4",100,240,400,"Low","AA nonstop is shorter; fewer fees than connecting","SL-FL","Both"],
    # row 4
    ["Lodging (A)","Coconut Bay Splash wing — all-inclusive, 10 nights","$650–950/night","10 nights",6500,7800,9500,"Medium","cbayresort.com; 2026 rates TBC; estimate based on historical July AI pricing","SL-LO","Option A only"],
    # row 5
    ["Lodging (B)","Soufrière villa 2BR with kitchen — 6 nights","$300–450/night","6 nights",1800,2280,2700,"Medium","Stonefield Villa Resort / Fond Doux / private rental (Airbnb/VRBO)","SL-LO","Option B only"],
    # row 6
    ["Lodging (B)","Bay Gardens Beach Resort, Rodney Bay — 4 nights","$250–380/night","4 nights",1000,1160,1520,"Medium","Bay Gardens direct booking","SL-LO","Option B only"],
    # row 7
    ["Transfers (A)","UVF → Coconut Bay (free resort shuttle)","$0","—",0,0,0,"High","Coconut Bay direct confirmation","SL-TR","Option A only"],
    # row 8
    ["Transfers (A)","Excursion day taxis (not resort-included)","Estimate","—",0,120,200,"Low","Distance from Vieux Fort adds cost","SL-TR","Option A only"],
    # row 9
    ["Transfers (B)","UVF → Soufrière (private taxi for 4)","$90–100","1",90,95,110,"High","Representative local operator pricing","SL-TR","Option B only"],
    # row 10
    ["Transfers (B)","Soufrière → Rodney Bay (taxi or water taxi)","$100–180","1",100,130,180,"Medium","Water taxi is scenic upgrade; road taxi is practical","SL-TR","Option B only"],
    # row 11
    ["Transfers (B)","Rodney Bay → UVF (departure)","$100–130","1",100,115,130,"High","Representative local operator pricing","SL-TR","Option B only"],
    # row 12
    ["Transfers (B)","Local taxis throughout trip","Estimate","—",200,300,450,"Low","Island roads require taxis; no public transit grid","SL-TR","Option B only"],
    # row 13
    ["Activities","Gros Piton hike (trail fee + mandatory guide)","$50/pp","4",200,215,240,"High","St. Lucia National Trust; guide required by law","SL-AC","Both"],
    # row 14
    ["Activities","Soufrière highlights (Sulphur Springs + Toraille + Diamond)","$90–110/pp","4",280,400,560,"Medium","Direct operator and St. Lucia tourism authority pricing","SL-AC","Both"],
    # row 15
    ["Activities","Catamaran full-day cruise (Sea Spray Tout Bagay or equiv.)","$140/adult; $80/child","4",360,440,560,"Medium","Sea Spray Cruises direct pricing","SL-AC","Both"],
    # row 16
    ["Activities","Rainforest zip-line (Rainforest Adventures or Morne Coubaril)","$75–95/pp","4",280,360,480,"Medium","Operator pricing; min age 8; max 290 lbs","SL-AC","Both"],
    # row 17
    ["Activities","Tet Paul Nature Trail","$10/pp","4",40,40,40,"High","Tet Paul official site","SL-AC","Both"],
    # row 18
    ["Activities","Pigeon Island National Park","$11/pp","4",0,44,44,"High","St. Lucia National Trust","SL-AC","Both"],
    # row 19
    ["Activities","Anse La Raye Fish Fry transport","$30–40/trip","2 trips",0,70,90,"Low","Driver estimate","SL-TR","Both"],
    # row 20
    ["Activities (opt.)","Hotel Chocolat / Boucan estate tour","$95/pp","4",0,0,380,"Low","Hotel Chocolat direct","SL-AC","Both"],
    # row 21
    ["Activities (opt.)","Anse Chastanet snorkel transfer","$40/pp","4",0,0,160,"Low","Optional upgrade","SL-AC","Both"],
    # row 22
    ["Food (A)","Off-resort meals at all-inclusive (2–3 outings)","$60/meal × family","3 meals",150,300,600,"Low","Most meals included at AI resort","SL-FD","Option A only"],
    # row 23
    ["Food (B)","10 days × 4 people","$50/$75/$110 pp/day","40 pp-days",2000,3000,4400,"Medium","Cook breakfast in villa; local lunches $10–15/pp; dinners $30–50/pp","SL-FD","Option B only"],
    # row 24
    ["Tips & Misc","Guides, resort tips, incidentals","Estimate","—",100,200,350,"Low","Estimate","—","Both"],
    # row 25
    ["Travel Insurance","Hurricane/CFAR policy (strongly recommended for July)","$100/pp","4",300,450,600,"Medium","July = hurricane season; St. Lucia Holiday Guarantee also available","—","Both"],
]

d4 = {5: USD, 6: USD, 7: USD}
for i, row in enumerate(stlucia_costs):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d4)

# TOTAL A — row 26: SUMIF on col K for "Both" + "Option A only"
sumif_a_lo = '=SUMIF($K$2:$K$25,"Both",E2:E25)+SUMIF($K$2:$K$25,"Option A only",E2:E25)'
sumif_a_mi = '=SUMIF($K$2:$K$25,"Both",F2:F25)+SUMIF($K$2:$K$25,"Option A only",F2:F25)'
sumif_a_hi = '=SUMIF($K$2:$K$25,"Both",G2:G25)+SUMIF($K$2:$K$25,"Option A only",G2:G25)'
write_row(ws, 26,
    ["TOTAL — Option A (Coconut Bay All-Inclusive)", "", "", "",
     sumif_a_lo, sumif_a_mi, sumif_a_hi, "", "", "", "A"],
    bg=TOTAL_CLR, bold=True, num_fmt=d4)

# TOTAL B — row 27: SUMIF on col K for "Both" + "Option B only"
sumif_b_lo = '=SUMIF($K$2:$K$25,"Both",E2:E25)+SUMIF($K$2:$K$25,"Option B only",E2:E25)'
sumif_b_mi = '=SUMIF($K$2:$K$25,"Both",F2:F25)+SUMIF($K$2:$K$25,"Option B only",F2:F25)'
sumif_b_hi = '=SUMIF($K$2:$K$25,"Both",G2:G25)+SUMIF($K$2:$K$25,"Option B only",G2:G25)'
write_row(ws, 27,
    ["TOTAL — Option B (Villa + Bay Gardens)", "", "", "",
     sumif_b_lo, sumif_b_mi, sumif_b_hi, "", "", "", "B"],
    bg=TOTAL_CLR, bold=True, num_fmt=d4)

set_col_widths(ws, [18, 46, 24, 12, 12, 12, 12, 14, 52, 12, 14])
freeze(ws)

# ============================================================
# TAB 9 — Greece Option 2 Costs
# Row map (data rows 2-20, TOTAL row 21):
#   2-3:   Flights IAD-ATH, Bags & Seats
#   4-5:   Domestic flights ATH→CHQ, HER→ATH
#   6-8:   Lodging (West Crete 5n, East Crete/Agios Nikolaos 3n, Athens 1n)
#   9-10:  Rental car, Fuel & parking
#   11-18: Activities (8 rows)
#   19:    Food
#   20:    Travel Insurance
# Verified totals: Low $9,280 / Mid $13,005 / High $17,185
# ============================================================
ws = wb.create_sheet("Greece2_Costs")
ws.row_dimensions[1].height = 36
write_header(ws, 1, COST_COLS)

greece2_costs = [
    # row 2
    ["Flights","IAD→ATH round-trip economy × 4","$1,000–1,600/pp","4",4000,5200,6400,"Medium","Live IAD–ATH searches May 2026; Aegean/United/Lufthansa range","GR-FL"],
    # row 3
    ["Bags & Seats","Economy bags + seat selection","$50–140/pp RT","4",200,360,560,"Low","Carrier-dependent; assume 1 checked bag/pp","GR-FL"],
    # row 4
    ["Domestic Flight","Athens → Chania (ATH→CHQ; Aegean Airlines)","$70–120/pp","4",280,380,480,"High","Aegean Airlines; book 2–3 months ahead","GR2-DO"],
    # row 5
    ["Domestic Flight","Heraklion → Athens (HER→ATH; Aegean Airlines)","$70–120/pp","4",280,380,480,"High","Aegean Airlines; book 2–3 months ahead","GR2-DO"],
    # row 6
    ["Lodging","West Crete (Agia Marina/Stalos/Kalamaki) — 5 nights","$180–280/night","5",900,1150,1400,"Medium","Booking.com; beach-area family hotels with pool and parking","GR2-LO"],
    # row 7
    ["Lodging","East Crete — Agios Nikolaos mid-tier — 3 nights","$150–260/night","3",450,660,780,"Medium","Mid-tier hotel near harbor; Agios Nikolaos recommended over Elounda for value","GR2-LO"],
    # row 8
    ["Lodging","Athens — 1 night (Plaka/Koukaki area)","$150–280/night","1",150,200,280,"Medium","Central Athens hotel; Plaka or Koukaki walkable to Acropolis","GR2-LO"],
    # row 9
    ["Rental Car","Economy/compact 9 days CHQ pickup → HER return","$50–85/day","9",450,580,765,"Medium","Budget/Sixt/Europcar Crete; CHQ one-way to HER adds drop fee ~$50–100","GR2-CA"],
    # row 10
    ["Fuel & Parking","Crete driving 9 days (~700 km total)","Estimate","—",150,200,300,"Low","Crete petrol ~€1.90/L May 2026; minimal tolls; parking €5–10/day in cities","GR2-CA"],
    # row 11
    ["Activities","Knossos Palace (timed entry)","€15/adult; €8/youth","4",50,55,55,"High","Greek Ministry of Culture / hhticket.gr","GR2-AC"],
    # row 12
    ["Activities","Heraklion Archaeological Museum","€15/adult; €8/youth","4",50,55,55,"High","Official museum site; world-class Minoan collection","GR2-AC"],
    # row 13
    ["Activities","Acropolis timed entry Day 11 — FIRM","€30/adult; €15/youth","4",120,170,170,"High","hhticket.gr; book immediately — July 7:30 am slots sell out","GR2-AC"],
    # row 14
    ["Activities","Spinalonga boat + island entry","€15–25/pp RT + €8 entry","4",70,105,140,"High","Elounda–Spinalonga local operators; book morning boat","GR2-AC"],
    # row 15
    ["Activities","Imbros Gorge entry","€5/pp + taxi return from Sfakia","4",10,10,10,"High","Pay at trailhead; taxi return ~€15 total","GR2-AC"],
    # row 16
    ["Activities","Balos Lagoon (Kissamos ferry round-trip)","€18–25/pp","4",70,100,140,"Medium","Kissamos ferry operator; pre-book July; alternatively drive to overlook","GR2-AC"],
    # row 17
    ["Activities (opt.)","Inland village olive oil / food tasting","€20/pp","4",0,80,130,"Low","Optional; Cretan olive oil estates in Vamos / Gavalochori area","—"],
    # row 18
    ["Activities","Misc (Rethymno Fortezza, Sfakia taxi, beach incidentals, parking)","Estimate","—",50,120,220,"Low","Estimate","—"],
    # row 19
    ["Food & Dining","9 ground days × 4 people","$50/$80/$120 pp/day","36 pp-days",1800,2880,4320,"Medium","Crete tavernas competitive; Agios Nikolaos harbor restaurants mid-range","GR2-FD"],
    # row 20
    ["Travel Insurance","Trip protection (rental car + flight delays)","$80/pp","4",200,320,500,"Medium","Recommended; ensure rental car coverage included","—"],
]

for i, row in enumerate(greece2_costs):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d3)

# TOTAL row 21 — formulas sum rows 2:20
# Low=$9,280 / Mid=$13,005 / High=$17,185
write_row(ws, 21,
    ["TOTAL", "", "", "", "=SUM(E2:E20)", "=SUM(F2:F20)", "=SUM(G2:G20)", "", "", ""],
    bg=TOTAL_CLR, bold=True, num_fmt=d3)

set_col_widths(ws, [18, 46, 24, 10, 12, 12, 12, 14, 52, 12])
freeze(ws)

# ============================================================
# TAB 12 — Cost Comparison (Mid-Range)
# 7 columns: A=Category, B=Greece Opt1, C=Greece Opt2, D=Bavaria Opt1,
#            E=Bavaria Opt2, F=StLucia-A, G=StLucia-B
# All values pulled via cross-sheet formulas.
# Row map:
#   2: Flights + bags
#   3: Lodging
#   4: Ground transport / car / rail / ferries
#        BA2: rows 7-12 (rail+transit only; domestics not applicable)
#   5: Activities
#        BA1: rows 13-25; BA2: rows 13-24
#   6: Food       BA1: row 26; BA2: row 25
#   7: Insurance  BA1: row 27; BA2: row 26
#   8: TOTAL MID
#   9: Nights on ground (Summary C2:C7)
#  10: Cost per night mid
# ============================================================
ws = wb.create_sheet("Cost_Comparison")
ws.row_dimensions[1].height = 36
write_header(ws, 1, [
    "Category (Mid-Range)",
    "Greece — Opt 1 (Cyclades)",
    "Greece — Opt 2 (Crete)",
    "Bavaria — Opt 1 (Alpine)",
    "Bavaria — Opt 2 (Cities)",
    "St. Lucia (A) All-Incl.",
    "St. Lucia (B) Independent"
])

comp_rows = [
    # row 2: Flights + bags
    ["Flights + Bags",
     "=SUM(Greece_Costs!F2:F3)",
     "=SUM(Greece2_Costs!F2:F5)",
     "=SUM(Bavaria_Costs!F2:F3)",
     "=SUM(Bavaria2_Costs!F2:F3)",
     "=SUM(StLucia_Costs!F2:F3)",
     "=SUM(StLucia_Costs!F2:F3)"],
    # row 3: Lodging
    ["Lodging",
     "=SUM(Greece_Costs!F4:F7)",
     "=SUM(Greece2_Costs!F6:F8)",
     "=SUM(Bavaria_Costs!F4:F7)",
     "=SUM(Bavaria2_Costs!F4:F6)",
     "=StLucia_Costs!F4",
     "=SUM(StLucia_Costs!F5:F6)"],
    # row 4: Ground transport / car / rail / ferries
    ["Ground Transport / Ferries / Rail",
     "=SUM(Greece_Costs!F8:F12)",
     "=SUM(Greece2_Costs!F9:F10)",
     "=SUM(Bavaria_Costs!F8:F12)",
     "=SUM(Bavaria2_Costs!F7:F12)",
     "=SUM(StLucia_Costs!F7:F8)",
     "=SUM(StLucia_Costs!F9:F12)"],
    # row 5: Activities
    ["Activities",
     "=SUM(Greece_Costs!F13:F20)",
     "=SUM(Greece2_Costs!F11:F18)",
     "=SUM(Bavaria_Costs!F13:F25)",
     "=SUM(Bavaria2_Costs!F13:F24)",
     "=SUM(StLucia_Costs!F13:F21)",
     "=SUM(StLucia_Costs!F13:F21)"],
    # row 6: Food
    ["Food & Dining",
     "=Greece_Costs!F21",
     "=Greece2_Costs!F19",
     "=Bavaria_Costs!F26",
     "=Bavaria2_Costs!F25",
     "=StLucia_Costs!F22",
     "=StLucia_Costs!F23"],
    # row 7: Insurance & misc
    ["Insurance & Misc",
     "=Greece_Costs!F22",
     "=Greece2_Costs!F20",
     "=Bavaria_Costs!F27",
     "=Bavaria2_Costs!F26",
     "=SUM(StLucia_Costs!F24:F25)",
     "=SUM(StLucia_Costs!F24:F25)"],
]

d_comp = {2: USD, 3: USD, 4: USD, 5: USD, 6: USD, 7: USD}
for i, row in enumerate(comp_rows):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE, num_fmt=d_comp)

# TOTAL MID — row 8
write_row(ws, 8,
    ["TOTAL MID (Family of 4)",
     "=SUM(B2:B7)", "=SUM(C2:C7)", "=SUM(D2:D7)", "=SUM(E2:E7)", "=SUM(F2:F7)", "=SUM(G2:G7)"],
    bg=TOTAL_CLR, bold=True, num_fmt=d_comp)

# Nights on ground — row 9: single source of truth from Summary tab
# Summary rows: 2=Greece Opt1, 3=Greece Opt2, 4=Bavaria Opt1, 5=Bavaria Opt2, 6=StLucia-A, 7=StLucia-B
write_row(ws, 9,
    ["Nights on Ground",
     "=Summary!C2", "=Summary!C3", "=Summary!C4", "=Summary!C5", "=Summary!C6", "=Summary!C7"],
    bg=WHITE)

# Cost per night — row 10
write_row(ws, 10,
    ["Cost per Night on Ground (Mid)",
     "=B8/B9", "=C8/C9", "=D8/D9", "=E8/E9", "=F8/F9", "=G8/G9"],
    bg=HDR_LIGHT, num_fmt=d_comp)

set_col_widths(ws, [36, 20, 20, 20, 20, 24, 24])
freeze(ws)

# ============================================================
# TAB 11 — Source Register
# ============================================================
ws = wb.create_sheet("Source_Register")
ws.row_dimensions[1].height = 36
write_header(ws, 1, ["Source ID", "Destination", "Category", "Source Name",
                      "URL / Where to Find", "What It Supports", "Reliability", "Notes"])
sources = [
    # Greece Option 1 sources
    ["GR-FL","Greece Opt 1","Flights","United Airlines IAD–ATH seasonal nonstop","united.com","IAD–ATH nonstop routing and fare confirmation","High","Seasonal summer nonstop; verify July 2026 schedule"],
    ["GR-FL","Greece Opt 1","Flights","Lufthansa IAD–FRA–ATH","lufthansa.com","Connecting option; typically competitive in advance","High",""],
    ["GR-FL","Greece Opt 1","Flights","Google Flights / Kayak","flights.google.com / kayak.com","Fare market scanning and range validation","Medium","Fares highly volatile; use as range only"],
    ["GR-AC","Greece Opt 1","Attractions","Greek Ministry of Culture / hhticket.gr","hhticket.gr","Acropolis timed-entry booking; official pricing (€30 flat 2026)","High","Book timed entry immediately; July slots sell out weeks ahead"],
    ["GR-AC","Greece Opt 1","Attractions","Acropolis Museum official","theacropolismuseum.gr","Museum hours; admission pricing (€15 adult)","High",""],
    ["GR-FE","Greece Opt 1","Ferries","Blue Star Ferries","bluestarferries.com","Piraeus–Naxos schedules; pricing; stability in meltemi winds","High","Blue Star far more wind-stable than Seajets; recommended for Piraeus–Naxos leg"],
    ["GR-FE","Greece Opt 1","Ferries","Seajets","seajets.com","Fast-ferry schedules Naxos–Santorini; pricing","High","Weather-sensitive; check before travel"],
    ["GR-FE","Greece Opt 1","Ferries","Ferryhopper (aggregator)","ferryhopper.com","Ferry booking aggregator; compares operators","Medium-High","Good for comparing Blue Star vs. Seajets options"],
    ["GR-DF","Greece Opt 1","Domestic Flights","Aegean Airlines","aegeanair.com","Santorini → Athens domestic; 5–6 daily; ~50 min","High","Recommended for buffer-day return to Athens"],
    ["GR-LO","Greece Opt 1","Lodging","Booking.com / Expedia / direct hotels","booking.com","Athens, Naxos, Santorini family room rates","Medium","July peak; Santorini caldera-view hotels sell out 6+ months ahead"],
    ["GR-TR","Greece Opt 1","Transfers","Proastiakos Athens Metro (airport express)","stasy.gr","Athens airport express train; fares; under-18 half price","High",""],
    ["GR-FD","Greece Opt 1","Food","Greek National Tourism Org / Santorini Dave","visitgreece.gr","Dining cost benchmarks for Athens, Naxos, Santorini","Medium","Estimate; confirm on arrival"],
    # Greece Option 2 sources
    ["GR2-DO","Greece Opt 2","Domestic Flights","Aegean Airlines ATH–CHQ / HER–ATH","aegeanair.com","Both Crete domestic legs; ~50 min each; book 2–3 months ahead","High","Book together with main IAD–ATH international itinerary"],
    ["GR2-LO","Greece Opt 2","Lodging","Booking.com / direct Crete hotels","booking.com","West Crete (Agia Marina/Stalos) and East Crete (Agios Nikolaos) family hotel rates","Medium","Book Agios Nikolaos mid-tier hotel 4–5 months ahead for July"],
    ["GR2-CA","Greece Opt 2","Car Rental","Budget / Sixt / Europcar Crete","budget.com / sixt.com","Economy/compact rental CHQ pickup one-way to HER; 9 days","Medium","One-way drop fee varies; compare operators directly; fuel full-to-full"],
    ["GR2-AC","Greece Opt 2","Attractions","Greek Ministry of Culture / hhticket.gr — Knossos & Heraklion Museum","hhticket.gr","Knossos Palace timed entry and Heraklion Archaeological Museum pricing","High","Knossos timed entry recommended in July; museum is world-class"],
    ["GR2-AC","Greece Opt 2","Attractions","hhticket.gr — Acropolis timed entry","hhticket.gr","Acropolis Day 11 firm activity; €30 adult flat rate 2026","High","Book immediately; July 7:30 am slots sell out weeks ahead"],
    ["GR2-AC","Greece Opt 2","Attractions","Elounda–Spinalonga local boat operators","elounda.gr / local dock","Spinalonga round-trip boat and island entry; morning boats recommended","High","Multiple operators at Elounda dock; pre-book July"],
    ["GR2-AC","Greece Opt 2","Attractions","Kissamos Ferry (Balos Lagoon)","kissamos-ferries.gr","Kissamos–Balos return ferry; seasonal May–Oct; July books up","Medium","Pre-book ferry; alternatively drive to Balos overlook for free"],
    ["GR2-FD","Greece Opt 2","Food","Crete dining benchmarks","visitcrete.gr / local restaurant guides","Cretan taverna food cost benchmarks; Agios Nikolaos harbor restaurants","Medium","Crete generally good value vs. Santorini; harbor slightly premium"],
    # Bavaria sources
    ["BA-FL","Bavaria","Flights","Lufthansa IAD–MUC nonstop","lufthansa.com","Direct nonstop service; ~8h 10m","High","Also check United IAD–MUC"],
    ["BA-FL","Bavaria","Flights","Icelandair via KEF","icelandair.com","Budget routing option; adds ~2–3 hrs","Medium","Good for low scenario pricing"],
    ["BA-RT","Bavaria","Rail","Deutsche Bahn Bayern-Ticket","bahn.de","Bayern-Ticket pricing; validity; free-child rule (ages 6–14 free)","High","14yo rides free; 17yo is adult rate; €49/family of 4 per day"],
    ["BA-RT","Bavaria","Rail","ÖBB Railjet Salzburg–Munich","oebb.at","Salzburg → Munich final morning; 1h 28m; seat reservation required","High","Book 3–4 weeks ahead; reservation ~$30/person"],
    ["BA-AC","Bavaria","Attractions","Zugspitze official","zugspitze.de","Summer cable car operations; family/youth pricing; webcam for conditions","High","Check webcam morning of; youth pricing applies for 17yo"],
    ["BA-AC","Bavaria","Attractions","Karwendelbahn official","karwendelbahn.de","Mittenwald cable car; schedule; pricing; free under 15","High","14yo rides free"],
    ["BA-AC","Bavaria","Attractions","Bayerische Seenschifffahrt (Königssee)","seenschifffahrt.de","Electric boat schedule; pricing; adult/youth rates","High","Take first morning boat; seats go fast"],
    ["BA-AC","Bavaria","Attractions","Kehlstein (Eagle's Nest)","kehlsteinhaus.de","Bus/elevator access; seasonal operation; ticket pricing","High","Cash/credit card at top; operates May–Oct"],
    ["BA-AC","Bavaria","Attractions","Deutsches Museum","deutsches-museum.de","Admission pricing; hours; teen-appeal content","High",""],
    ["BA-AC","Bavaria","Attractions","Festung Hohensalzburg","salzburg-burgen.at","Fortress funicular pricing; adult/youth rates","High",""],
    ["BA-LO","Bavaria","Lodging","Booking.com / direct hotels","booking.com","Munich / Mittenwald / Berchtesgaden / Salzburg family room rates","Medium","Salzburg: book 5–6 months ahead due to Festspiele"],
    ["BA-FD","Bavaria","Food","Munich / Bavaria general tourism","munich.travel","Food cost benchmarks; beer garden pricing","Medium","Bavaria moderately priced; bakery + beer garden strategy keeps costs low"],
    ["BA2-AC","Bavaria Opt 2","Attractions","Neuschwanstein / Hohenschwangau official tickets","tickets.hohenschwangau.de","Neuschwanstein timed interior tour; under 18 free with guardian; July availability","High","Book immediately upon confirming trip; July sells out weeks ahead"],
    ["BA2-AC","Bavaria Opt 2","Attractions","Schloss Nymphenburg official","schloss-nymphenburg.de","Palace entry pricing; grounds and canals free; summer café on-site","High","Good half-day activity; grounds free even without palace entry"],
    ["BA2-AC","Bavaria Opt 2","Attractions","Hallstatt tourism / ÖBB regional rail","hallstatt.net / oebb.at","Hallstatt day trip logistics from Salzburg; train + bus combo; Salzwelten mine museum","Medium","Regional rail Salzburg–Attnang-Puchheim–Hallstatt; arrive early to beat July tour buses"],
    # St. Lucia sources
    ["SL-FL","St. Lucia","Flights","American Airlines PHL–UVF","aa.com","Weekly nonstop PHL–UVF; ~4h 45m; fare range","High","Verify exact July 2026 schedule; check early for PHL nonstop availability"],
    ["SL-LO","St. Lucia","Lodging","Coconut Bay Resort","cbayresort.com","All-inclusive family Splash wing; rates; inclusions","Medium","2026 July rates not yet published; get direct quote"],
    ["SL-LO","St. Lucia","Lodging","Bay Gardens Beach Resort","baygardensresorts.com","Rodney Bay family rooms; nightly rates","High",""],
    ["SL-LO","St. Lucia","Lodging","Stonefield Villa Resort","stonefieldvillas.com","Soufrière villa option; 2BR; Piton views","High",""],
    ["SL-AC","St. Lucia","Activities","St. Lucia National Trust (Gros Piton)","slunatrust.org","Gros Piton trail fee ($50/pp); mandatory guide requirement","High","Guide is legally required; adds real value"],
    ["SL-AC","St. Lucia","Activities","Tet Paul Nature Trail","thetetpaulnaturetrail.com","Entry pricing; trail description; panoramic Piton views","High",""],
    ["SL-AC","St. Lucia","Activities","Sea Spray Cruises","seaspreystlucia.com","Tout Bagay full-day catamaran; pricing; schedule","High",""],
    ["SL-AC","St. Lucia","Activities","Rainforest Adventures St. Lucia","rainforestadventure.com","Aerial tram + zip-line; pricing; age/weight requirements","High","Min age 8; max 290 lbs"],
    ["SL-AC","St. Lucia","Activities","Diamond Botanical Gardens","diamondstlucia.com","Entry pricing; mineral baths; waterfall","High",""],
    ["SL-TR","St. Lucia","Transfers","Local taxi operators","stluciataxiservice.com","UVF transfer cost estimates; Soufrière / Rodney Bay rates","Medium","Get quotes directly on arrival; prices can vary"],
    ["SL-WX","St. Lucia","Weather","NOAA / National Hurricane Center","nhc.noaa.gov","July hurricane season framing; historical risk profile","High","Early July historically lower risk than Aug/Sep; buy CFAR travel insurance"],
    ["SL-WX","St. Lucia","Weather","Saint Lucia Tourism Authority","stlucia.org","July weather and travel season guidance","High",""],
    # Cross-destination
    ["FX","All","FX Reference","ECB EUR/USD reference rate (May 2026)","ecb.europa.eu","EUR to USD conversion: €1 = $1.18 used throughout","High","Verify current rate when booking; significant FX moves possible before July 2026"],
    ["US-DOS","All","Safety","U.S. Department of State","travel.state.gov","Travel advisories: Greece L1; Germany L1; Austria L1; St. Lucia L1","High","Verify before departure"],
]
for i, row in enumerate(sources):
    write_row(ws, i + 2, row, bg=HDR_LIGHT if i % 2 == 0 else WHITE)
set_col_widths(ws, [12, 14, 16, 34, 32, 46, 12, 44])
freeze(ws)

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\jimbu\Coding Projects\temp3\vacation_research_2026.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
