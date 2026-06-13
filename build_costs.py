#!/usr/bin/env python3
"""
build_costs.py — generates trip_costs.xlsx and updates the Cost Estimate
section in itinerary.md and itinerary.html.

Run: python build_costs.py
Requires: openpyxl  (pip install openpyxl)

To update a cost: edit the COSTS list below and re-run.
Totals are computed by Python — not the LLM, not Excel formulas.
The spreadsheet uses SUM() formulas as a double-check.
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent

# ── Cost line items ────────────────────────────────────────────────────────────
# (category, description, USD_amount, is_estimate)
#   is_estimate=False  confirmed/booked price
#   is_estimate=True   placeholder — research and update before trip

COSTS = [
    # FLIGHTS
    ("Flights",
     "UA 108 IAD→MUC Jul 11 + OS 41 VIE→IAD Jul 22 nonstop (4 people) — Conf. IKR3ZV",
     5422.52, False),

    # ACCOMMODATION
    ("Accommodation", "Munich — PRIME City Apt Adamstraße 4, 2 nights (Jul 12–14) — Conf. HM4ZPJHYB9", 782.08, False),
    ("Accommodation", "Königssee — Villa Alpenrausch, 3 nights (Jul 14–17) BOOKED", 1435.99, False),
    ("Accommodation", "Salzburg — Mozart House Gstättengasse 25, 2 nights (Jul 17–19) — Conf. HM4NSFM9J3", 1007.28, False),
    ("Accommodation", "Vienna — City Center Klimt's Balcony Suite, 3 nights (Jul 19–22) — Conf. HMT8SQRKMA", 1068.21, False),

    # TRANSPORT
    ("Transport", "Airport → city: MVV group day ticket (Jul 12)",              16.00, True),
    ("Transport", "Bayern-Ticket: Munich → Berchtesgaden (Jul 14, group)",      58.00, True),
    ("Transport", "Local buses Berchtesgaden area (Jul 15–16, ~€8/day)",        18.00, True),
    ("Transport", "Schönau → Salzburg: taxi to Freilassing + ÖBB (Jul 17)",    100.00, True),
    ("Transport", "ÖBB Railjet: Salzburg → Vienna (Jul 19, 4 people)",         120.00, True),
    ("Transport", "Vienna public transit (Jul 19–22)",                           28.00, True),
    ("Transport", "CAT Wien Mitte → VIE airport (Jul 22, 4 people)",            52.00, True),

    # ACTIVITIES
    ("Activities", "Mike's Bike Tours Munich (4 people) — €180.67 paid, Booking #354406124", 197.00, False),
    ("Activities", "Nymphenburg Palace combination ticket — €58 paid, Order #42616", 63.00, False),
    ("Activities", "Eagle's Nest Kehlstein bus (4 people)",                      80.00, True),
    ("Activities", "Königssee electric boat (4 people)",                         88.00, True),
    ("Activities", "Hohensalzburg fortress admission (4 people)",                66.00, True),
    ("Activities", "St. Peter Stiftskeller dinner (4 people)",                  220.00, True),
    ("Activities", "Schönbrunn Palace + maze (4 people)",                       132.00, True),
    ("Activities", "Upper Belvedere admission (4 people)",                       84.00, True),
    ("Activities", "Riesenrad / Prater (4 people)",                              64.00, True),

    # MEALS & INCIDENTALS
    ("Meals & Incidentals", "Meals (~$150/day × 11 days)",      1650.00, True),
    ("Meals & Incidentals", "Miscellaneous / buffer",             500.00, True),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def usd(n: float) -> str:
    return f"${n:,.0f}"

def get_subtotals() -> dict[str, float]:
    result: dict[str, float] = {}
    for cat, _, amt, _ in COSTS:
        result[cat] = result.get(cat, 0.0) + amt
    return result

def get_grand_total() -> float:
    return sum(amt for _, _, amt, _ in COSTS)

def grouped_costs() -> list[tuple[str, list, float]]:
    """Returns [(category, items, subtotal), ...] preserving order."""
    groups: list[tuple[str, list, float]] = []
    for cat, item, amt, est in COSTS:
        if groups and groups[-1][0] == cat:
            groups[-1][1].append((cat, item, amt, est))
            groups[-1] = (groups[-1][0], groups[-1][1], groups[-1][2] + amt)
        else:
            groups.append((cat, [(cat, item, amt, est)], amt))
    return groups

# ── Spreadsheet ───────────────────────────────────────────────────────────────

def build_spreadsheet() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trip Costs"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 16

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    h_fill = fill("2C5F8A")
    c_fill = fill("D5E8F5")
    t_fill = fill("2C5F8A")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    c_font = Font(bold=True, size=10)
    t_font = Font(bold=True, color="FFFFFF", size=11)
    money_fmt = '"$"#,##0.00'
    center = Alignment(horizontal="center")

    # Header row
    ws.append(["Category", "Item", "Amount (USD)"])
    for col in range(1, 4):
        cell = ws.cell(1, col)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center

    row = 1
    subtotal_rows: list[int] = []
    for cat, items, cat_sub in grouped_costs():
        # Category header row — subtotal formula goes here in col C
        row += 1
        cat_header_row = row
        ws.cell(row, 1, cat).font = c_font
        for col in range(1, 4):
            ws.cell(row, col).fill = c_fill

        # Item rows
        data_start = row + 1
        for _, item, amt, est in items:
            row += 1
            label = item + (" *" if est else "")
            item_cell = ws.cell(row, 2, label)
            item_cell.font = Font(italic=True) if est else Font()
            amt_cell = ws.cell(row, 3, amt)
            amt_cell.number_format = money_fmt
        data_end = row

        # Subtotal on the category header row
        sub_cell = ws.cell(cat_header_row, 3)
        sub_cell.value = f"=SUM(C{data_start}:C{data_end})"
        sub_cell.number_format = money_fmt
        sub_cell.font = c_font
        subtotal_rows.append(cat_header_row)

    # Blank row then TOTAL (sums only subtotal cells — no double-counting)
    row += 2
    ws.cell(row, 1, "TOTAL").font = t_font
    ws.cell(row, 1).fill = t_fill
    ws.cell(row, 1).alignment = Alignment(horizontal="right")
    ws.cell(row, 2).fill = t_fill
    total_cell = ws.cell(row, 3)
    total_cell.value = "=" + "+".join(f"C{r}" for r in subtotal_rows)
    total_cell.number_format = money_fmt
    total_cell.font = t_font
    total_cell.fill = t_fill

    row += 2
    ws.cell(row, 1, "* = estimated").font = Font(italic=True, color="888888")

    out = BASE / "trip_costs.xlsx"
    wb.save(out)
    print(f"  Wrote {out.name}")

# ── Sentinel markers (same string used in both MD and HTML) ───────────────────

MARKER_START = "<!-- COSTS_START -->"
MARKER_END   = "<!-- COSTS_END -->"

# ── Markdown section ──────────────────────────────────────────────────────────

def build_md_section() -> str:
    total = get_grand_total()
    lines = [
        "## Cost Estimate",
        "",
        "*Amounts in USD. Rows marked \\* are estimates — confirm before trip. "
        "Total computed by `build_costs.py`; source data: `trip_costs.xlsx`.*",
        "",
        "| Category | Item | Amount |",
        "|---|---|---|",
    ]
    for cat, items, cat_sub in grouped_costs():
        lines.append(f"| **{cat}** | | **{usd(cat_sub)}** |")
        for _, item, amt, est in items:
            label = item + (" \\*" if est else "")
            lines.append(f"| | {label} | {usd(amt)} |")

    lines += [
        "| | | |",
        f"| **TOTAL** | | **{usd(total)}** |",
    ]
    return "\n".join(lines)

def update_md() -> None:
    path = BASE / "itinerary.md"
    text = path.read_text(encoding="utf-8")
    block = f"{MARKER_START}\n{build_md_section()}\n{MARKER_END}"

    if MARKER_START in text:
        pattern = re.escape(MARKER_START) + r".*?" + re.escape(MARKER_END)
        text = re.sub(pattern, block, text, flags=re.DOTALL)
    else:
        # First run: insert before "## To Add"
        text = text.replace("## To Add", f"{block}\n\n---\n\n## To Add")

    path.write_text(text, encoding="utf-8")
    print(f"  Updated itinerary.md")

# ── HTML section ──────────────────────────────────────────────────────────────

COST_CSS = """\
    .cost-table { width: 100%; border-collapse: collapse; font-size: 0.93em; margin-top: 10px; }
    .cost-table th { background: #2c5f8a; color: #fff; padding: 7px 10px; text-align: left; }
    .cost-table td { padding: 5px 10px; border-bottom: 1px solid #e8e4dc; vertical-align: top; }
    .cost-cat-header td { background: #d5e8f5; font-weight: bold; padding: 6px 10px; }
    .cost-est td:nth-child(2) { font-style: italic; color: #555; }
    .cost-amt { text-align: right !important; font-variant-numeric: tabular-nums; white-space: nowrap; }
.cost-total-row td { background: #2c5f8a; color: #fff; font-weight: bold; padding: 8px 10px; border: none; }"""

def build_html_section() -> str:
    total = get_grand_total()
    rows: list[str] = []
    for cat, items, cat_sub in grouped_costs():
        rows.append(
            f'        <tr class="cost-cat-header">'
            f'<td colspan="2">{cat}</td>'
            f'<td class="cost-amt">{usd(cat_sub)}</td></tr>'
        )
        for _, item, amt, est in items:
            cls = ' class="cost-est"' if est else ""
            label = item + (" *" if est else "")
            rows.append(
                f'        <tr{cls}><td style="width:2em"></td>'
                f'<td>{label}</td>'
                f'<td class="cost-amt">{usd(amt)}</td></tr>'
            )

    return (
        f"{MARKER_START}\n"
        f"  <section>\n"
        f"    <h2>Cost Estimate</h2>\n"
        f"    <p style=\"font-size:0.88em;color:#666;margin-bottom:10px\">"
        f"All amounts USD. <em>Italic rows (*) are estimates — confirm before trip.</em> "
        f"Total verified by <code>build_costs.py</code> · source: <code>trip_costs.xlsx</code>.</p>\n"
        f"    <table class=\"cost-table\">\n"
        f"      <thead><tr><th colspan=\"2\">Item</th><th>Amount</th></tr></thead>\n"
        f"      <tbody>\n"
        + "\n".join(rows) + "\n"
        f"        <tr class=\"cost-total-row\">"
        f'<td colspan="2">TOTAL</td>'
        f'<td class="cost-amt">{usd(total)}</td></tr>\n'
        f"      </tbody>\n"
        f"    </table>\n"
        f"  </section>\n"
        f"{MARKER_END}"
    )

def update_html() -> None:
    path = BASE / "itinerary.html"
    text = path.read_text(encoding="utf-8")

    # Inject CSS once
    if ".cost-table" not in text:
        text = text.replace("</style>", COST_CSS + "\n  </style>")

    block = build_html_section()
    if MARKER_START in text:
        pattern = re.escape(MARKER_START) + r".*?" + re.escape(MARKER_END)
        text = re.sub(pattern, block, text, flags=re.DOTALL)
    else:
        # First run: insert before <!-- PHOTO GALLERY -->
        text = text.replace("  <!-- PHOTO GALLERY -->", f"{block}\n\n  <!-- PHOTO GALLERY -->")

    path.write_text(text, encoding="utf-8")
    print(f"  Updated itinerary.html")

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Building cost estimate…")
    build_spreadsheet()
    update_md()
    update_html()

    print("\nSummary:")
    subs = get_subtotals()
    for cat, amt in subs.items():
        print(f"  {cat:<26}  {usd(amt):>9}")
    print(f"  {'-' * 38}")
    print(f"  {'TOTAL':<26}  {usd(get_grand_total()):>9}")
